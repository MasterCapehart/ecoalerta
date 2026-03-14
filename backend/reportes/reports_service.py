"""
Servicio para generar reportes estadísticos y exportación
"""
import logging
from django.db.models import Count, Avg, Q, F
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger('reportes')


class ReportsService:
    """Servicio para generar reportes estadísticos"""
    
    @staticmethod
    def get_advanced_statistics(fecha_desde=None, fecha_hasta=None):
        """
        Obtiene estadísticas avanzadas de reportes
        """
        from .models import Reporte
        
        if fecha_desde is None:
            fecha_desde = timezone.now() - timedelta(days=30)
        if fecha_hasta is None:
            fecha_hasta = timezone.now()
        
        queryset = Reporte.objects.filter(
            fecha_creacion__gte=fecha_desde,
            fecha_creacion__lte=fecha_hasta
        )
        
        stats = {
            'periodo': {
                'desde': fecha_desde,
                'hasta': fecha_hasta
            },
            'totales': {
                'total': queryset.count(),
                'nuevos': queryset.filter(estado='nuevo').count(),
                'en_proceso': queryset.filter(estado='proceso').count(),
                'resueltos': queryset.filter(estado='resuelto').count(),
                'cerrados': queryset.filter(estado='cerrado').count(),
            },
            'por_categoria': list(
                queryset.values('categoria__nombre')
                .annotate(total=Count('id'))
                .order_by('-total')
            ),
            'por_prioridad': list(
                queryset.values('prioridad')
                .annotate(total=Count('id'))
                .order_by('-total')
            ),
            'por_dia': list(
                queryset.extra(
                    select={'dia': "DATE(fecha_creacion)"}
                ).values('dia')
                .annotate(total=Count('id'))
                .order_by('dia')
            ),
            'tiempo_promedio_resolucion': ReportsService._get_average_resolution_time(queryset),
            'tasa_resolucion': ReportsService._get_resolution_rate(queryset),
            'top_inspectores': ReportsService._get_top_inspectores(queryset),
        }
        
        return stats
    
    @staticmethod
    def _get_average_resolution_time(queryset):
        """Calcula tiempo promedio de resolución"""
        resueltos = queryset.filter(
            estado='resuelto',
            tiempo_resolucion_horas__isnull=False
        )
        
        if not resueltos.exists():
            return None
        
        promedio = resueltos.aggregate(
            avg=Avg('tiempo_resolucion_horas')
        )['avg']
        
        return promedio
    
    @staticmethod
    def _get_resolution_rate(queryset):
        """Calcula tasa de resolución"""
        total = queryset.count()
        if total == 0:
            return 0.0
        
        resueltos = queryset.filter(estado__in=['resuelto', 'cerrado']).count()
        return (resueltos / total) * 100
    
    @staticmethod
    def _get_top_inspectores(queryset):
        """Obtiene top inspectores por reportes resueltos"""
        from .models import Reporte
        
        return list(
            Reporte.objects.filter(
                id__in=queryset.values_list('id', flat=True),
                asignado_a__isnull=False,
                estado='resuelto'
            )
            .values('asignado_a__username')
            .annotate(
                total_resueltos=Count('id'),
                tiempo_promedio=Avg('tiempo_resolucion_horas')
            )
            .order_by('-total_resueltos')[:10]
        )
    
    @staticmethod
    def export_to_pdf(stats, output_stream):
        """
        Exporta estadísticas a PDF
        """
        doc = SimpleDocTemplate(output_stream, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph("Reporte Estadístico - EcoAlerta", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Período
        periodo_text = f"Período: {stats['periodo']['desde'].strftime('%d/%m/%Y')} - {stats['periodo']['hasta'].strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periodo_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Totales
        data = [
            ['Estado', 'Cantidad'],
            ['Total', str(stats['totales']['total'])],
            ['Nuevos', str(stats['totales']['nuevos'])],
            ['En Proceso', str(stats['totales']['en_proceso'])],
            ['Resueltos', str(stats['totales']['resueltos'])],
            ['Cerrados', str(stats['totales']['cerrados'])],
        ]
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Por categoría
        if stats['por_categoria']:
            elements.append(Paragraph("Reportes por Categoría", styles['Heading2']))
            cat_data = [['Categoría', 'Cantidad']]
            for item in stats['por_categoria']:
                cat_data.append([
                    item['categoria__nombre'] or 'Sin categoría',
                    str(item['total'])
                ])
            
            cat_table = Table(cat_data)
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(cat_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Métricas
        if stats['tiempo_promedio_resolucion']:
            tiempo_text = f"Tiempo promedio de resolución: {stats['tiempo_promedio_resolucion']:.2f} horas"
            elements.append(Paragraph(tiempo_text, styles['Normal']))
        
        tasa_text = f"Tasa de resolución: {stats['tasa_resolucion']:.2f}%"
        elements.append(Paragraph(tasa_text, styles['Normal']))
        
        doc.build(elements)
        return output_stream
    
    @staticmethod
    def export_to_excel(stats, output_stream):
        """
        Exporta estadísticas a Excel con gráficos
        """
        wb = openpyxl.Workbook()
        
        # Hoja de resumen
        ws = wb.active
        ws.title = "Resumen"
        
        ws['A1'] = "Reporte Estadístico - EcoAlerta"
        ws['A2'] = f"Período: {stats['periodo']['desde'].strftime('%d/%m/%Y')} - {stats['periodo']['hasta'].strftime('%d/%m/%Y')}"
        
        # Totales
        row = 4
        ws['A4'] = "Estado"
        ws['B4'] = "Cantidad"
        
        row = 5
        ws[f'A{row}'] = "Total"
        ws[f'B{row}'] = stats['totales']['total']
        row += 1
        ws[f'A{row}'] = "Nuevos"
        ws[f'B{row}'] = stats['totales']['nuevos']
        row += 1
        ws[f'A{row}'] = "En Proceso"
        ws[f'B{row}'] = stats['totales']['en_proceso']
        row += 1
        ws[f'A{row}'] = "Resueltos"
        ws[f'B{row}'] = stats['totales']['resueltos']
        row += 1
        ws[f'A{row}'] = "Cerrados"
        ws[f'B{row}'] = stats['totales']['cerrados']
        
        # Gráfico de barras
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Reportes por Estado"
        chart.y_axis.title = "Cantidad"
        chart.x_axis.title = "Estado"
        
        data = Reference(ws, min_col=2, min_row=4, max_row=row)
        cats = Reference(ws, min_col=1, min_row=5, max_row=row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 15
        
        ws.add_chart(chart, "D4")
        
        # Hoja por categoría
        if stats['por_categoria']:
            ws2 = wb.create_sheet("Por Categoría")
            ws2['A1'] = "Categoría"
            ws2['B1'] = "Cantidad"
            
            row = 2
            for item in stats['por_categoria']:
                ws2[f'A{row}'] = item['categoria__nombre'] or 'Sin categoría'
                ws2[f'B{row}'] = item['total']
                row += 1
        
        wb.save(output_stream)
        return output_stream

